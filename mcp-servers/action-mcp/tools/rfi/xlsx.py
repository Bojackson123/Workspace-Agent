"""Spreadsheet (.xlsx) structure dump and answer write-back."""

from __future__ import annotations

import io

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

from ._limits import _MAX_CELLS_PER_SHEET, _MAX_VALUE_LEN
from ._markdown import _md_to_plain


def _dump_xlsx(data: bytes) -> dict:
    """Dump every non-empty cell of every sheet with its real A1 address."""
    wb = load_workbook(io.BytesIO(data), data_only=True)
    sheets: list[dict] = []
    for ws in wb.worksheets:
        cells: list[dict] = []
        truncated = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value).strip()
                if not val:
                    continue
                cells.append({
                    "addr": f"{ws.title}!{cell.coordinate}",
                    "value": val[:_MAX_VALUE_LEN],
                })
            if len(cells) >= _MAX_CELLS_PER_SHEET:
                truncated = True
                break
        sheets.append({
            "sheet": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "cells": cells[:_MAX_CELLS_PER_SHEET],
            "truncated": truncated,
        })
    wb.close()
    return {"kind": "xlsx", "sheets": sheets}


def _fill_xlsx(data: bytes, answers: list[dict]) -> bytes:
    """Write answers into a workbook (formulas/formatting preserved) → bytes."""
    wb = load_workbook(io.BytesIO(data))  # data_only=False keeps formulas
    for ans in answers:
        location = ans.get("location") or ""
        sheet, _, coord = location.rpartition("!")
        if not sheet or sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        try:
            row, col = coordinate_to_tuple(coord)
        except Exception:  # noqa: BLE001
            continue
        ws.cell(row=row, column=col, value=_md_to_plain(ans.get("answer", "")))
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
